import shutil
import tkinter as tk
import os
from tkinter.filedialog import askopenfile
import tkinter.messagebox as msgbox
import openpyxl
import time
import customtkinter as ctk
from pathlib import Path

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")


class MainWindow(ctk.CTk):
    WIDTH = 525
    HEIGHT = 200

    def __init__(self):
        super().__init__()
        self.fontsize = ("Arial", 12)
        self.geometry(f"{MainWindow.WIDTH}x{MainWindow.HEIGHT}")
        self.resizable(False, False)
        self.title("Automated Drawing Retrieval Program")
        self.label1 = ctk.CTkLabel(self, text="Drawing list")
        self.label2 = ctk.CTkLabel(self, text="Source Location")
        self.label3 = ctk.CTkLabel(self, text="Output location")

        self.sourcelink = ctk.CTkEntry(width=250, border_width=2)
        self.destlink = ctk.CTkEntry(width=250, border_width=2)

        self.varlabel_text = tk.StringVar()
        self.varlabel_text.set("No drawings found yet")
        self.varlabel = ctk.CTkLabel(self, textvariable=self.varlabel_text)

        validatebutton = ctk.CTkButton(self, text="Validate", command=self.validatelink, width=20)
        self.transferbutton = ctk.CTkButton(self, text="Transfer", command=self.transfer, width=20)
        self.transferbutton['state'] = 'disable'
        browsebutton = ctk.CTkButton(self, text="Browse", command=self.browse, width=20)
        exitbutton = ctk.CTkButton(self, text="Exit", command=self.exit, width=20)

        self.bind('<Escape>', lambda e: exit())

        self.label1.grid(row=1, column=1, pady=10, padx=10, sticky='w')
        self.varlabel.grid(row=1, column=2, pady=10, padx=10)
        browsebutton.grid(row=1, column=3, pady=10, padx=10)

        self.label2.grid(row=2, column=1, pady=10, padx=10, sticky='w')
        self.sourcelink.grid(row=2, column=2, pady=10, padx=10)

        self.label3.grid(row=3, column=1, pady=10, padx=10, sticky='w')
        self.destlink.grid(row=3, column=2, pady=10, padx=10)

        validatebutton.grid(row=4, column=1, pady=10, padx=10)
        self.transferbutton.grid(row=4, column=2, pady=10, padx=10)
        exitbutton.grid(row=4, column=3, pady=10, padx=10)

        self.newlistofdrawings = []
        self.dwgexcel = ""

        deflink = "C:\\CITEC_EPDM\\2D-CAD\\Mech_dwg"
        self.sourcelink.insert(tk.END, deflink)

    def validatelink(self):
        getlink_1 = self.sourcelink.get()
        getlink_2 = self.destlink.get()
        if getlink_1 == "" or getlink_2 == "":
            msgbox.showwarning("Warning", "There is no link yet.")
        else:
            if not os.path.exists(getlink_1):
                msgbox.showerror("Error", f"This link is not found: {getlink_1}")
            elif not os.path.exists(getlink_2):
                msgbox.showerror("Error", f"This link is not found: {getlink_2}")
            else:
                msgbox.showinfo("Result", "Your link is valid!")
                self.transferbutton['state'] = 'active'

    def transfer(self):
        timeofstart = time.perf_counter()
        source = self.sourcelink.get()
        destination = self.destlink.get()
        ntdrawings = []
        successdrawings = []
        for i, drawing in enumerate(self.newlistofdrawings, start=0):
            drawing = self.newlistofdrawings[i]
            for rts, direcs, files in os.walk(source):
                for dwg in files:
                    name, ext = os.path.splitext(dwg)
                    if name == drawing:
                        os.chdir(rts)
                        if dwg not in os.listdir(destination):
                            shutil.copy2(dwg, destination)
                            successdrawings.append(drawing)

        for dwg in self.newlistofdrawings:
            if dwg not in successdrawings:
                ntdrawings.append(dwg)

        timeofend = time.perf_counter()
        runtime = timeofend - timeofstart

        ntdrawings = sorted(ntdrawings)

        wb = openpyxl.load_workbook(self.dwgexcel, data_only=True)
        ws = wb["Sheet1"]

        for dwg_rev in ntdrawings:
            dwg, rev = dwg_rev.split('_')
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
                if row[0].value == dwg:
                    rownum = int(row[0].row)
                    ws.cell(row=rownum, column=3).value = "Not found"
                    wb.save(self.dwgexcel)

        if len(ntdrawings) != 0:
            msgbox.showinfo("Missing drawings", f"Drawings not transferred: {ntdrawings}")
            msgbox.showinfo("Time taken is...", f"Your transfer took {round(runtime, 2)} seconds.")
            openfile = Path(self.dwgexcel).resolve()
            os.system(f'start excel.exe "{openfile}"')
        else:
            msgbox.showinfo("Time taken is...", f"Your transfer took {round(runtime, 2)} seconds.")

    def browse(self):
        emptydwglist = []
        self.transferbutton['state'] = 'disable'
        self.varlabel_text.set("Loading list of drawings")
        drawinglist = askopenfile(parent=self, title='Choose the file', mode='rb', filetypes=[("Excel file", "*.xlsx")])
        if drawinglist:
            self.dwgexcel = drawinglist.name
            workbook = openpyxl.load_workbook(drawinglist, data_only=True)
            sheet = workbook.active
            maxrow = sheet.max_row
            listofdrawings = []
            listofrevisions = []
            for i in range(1, maxrow + 1):
                drawingnumber = sheet.cell(row=i, column=1)
                listofdrawings.append(drawingnumber.value)
                revision = sheet.cell(row=i, column=2)
                listofrevisions.append(revision.value)

            for i in range(len(listofdrawings)):
                a = str(listofdrawings[i])
                b = str(listofrevisions[i])
                c = a + '_rev' + b
                if c != "Nonerev_None":
                    emptydwglist.append(c)

        emptydwglist = list(set(emptydwglist))

        if len(emptydwglist) > 0:
            self.varlabel_text.set(f"There are {len(emptydwglist)} drawings found.")
            msgbox.showinfo("Reminder", "Make sure the file is closed in Excel.")
        else:
            self.varlabel_text.set("Still no drawings found.")

        self.newlistofdrawings = emptydwglist

    def exit(self):
        self.destroy()


if __name__ == '__main__':
    main = MainWindow()
    main.mainloop()

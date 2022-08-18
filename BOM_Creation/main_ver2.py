import os
import sys
import tkinter as tk
import tkinter.messagebox as msgbox
from datetime import date
from pathlib import Path
from tkinter import ttk, Menu, simpledialog
from tkinter.filedialog import askopenfile

import openpyxl as openpyxl
from PIL import ImageTk, Image

from sql_query import query_stockcode, create_db


class MainWindow(tk.Toplevel):
    # inputname from class App for passing in new file name
    inputname = ''
    # status for indicating user start with new or load
    status = 0
    revdict, stockcodelist, dwgnumlist = {}, [], []

    def __init__(self, parent):
        super().__init__(parent)

        # Create window, window size, title and positioning on screen
        self.title("BOM Creation")

        self.screenwidth = self.winfo_screenwidth()
        self.screenheight = self.winfo_screenheight()

        self.width = self.screenwidth / 2
        self.height = self.screenheight

        self.paddingw = (self.screenwidth / 2) - (self.width / 2)
        self.paddingh = (self.screenheight / 2) - (self.height / 2)
        self.geometry('%dx%d+%d+%d' % (self.width, self.height, 0, self.paddingh))

        # style & color
        ttk.Style().configure("TButton", padding=2, relief='GROOVE', background="#3366cc")

        # create input box for stock code typing and display the information by vlookup in database
        # confirmation button to add into table if displayed information is OK
        self.lbl_category = ttk.Label(self, text="Category:")
        self.entry_category = ttk.Entry(self, width=90)
        self.lbl_stockcode = ttk.Label(self, text="Stock Code:")
        self.entry_stockcode = ttk.Entry(self, width=90)
        self.lbl_quantity = ttk.Label(self, text="Quantity:")
        self.entry_quantity = ttk.Entry(self, width=90)

        self.showmsg = tk.StringVar()
        self.showmsg.set("")
        self.refresh_status = tk.StringVar()
        self.refresh_status.set("Refresh DB")
        self.lbl_message = ttk.Label(self, textvariable=self.showmsg)

        self.lbl_count = ttk.Label(self, text="Number of stock code in BOM:")
        self.count = tk.StringVar()
        self.count.set("0")
        self.lbl_countnumber = ttk.Label(self, textvariable=self.count)

        self.lbl_itemqty = ttk.Label(self, text="Total number of items in BOM:")
        self.number = tk.StringVar()
        self.number.set("0")
        self.lbl_itemqtynumber = ttk.Label(self, textvariable=self.number)

        self.btn_table_entry = ttk.Button(self, text="Confirm", command=self.inserttable, width=15)
        self.btn_remove = ttk.Button(self, text="Delete", command=self.removeitem, width=15)

        self.btn_refresh = ttk.Button(self, textvariable=self.refresh_status, command=self.refresh_database, width=15)
        self.btn_refresh['state'] = 'disable'
        self.btn_export = ttk.Button(self, text="Export", command=self.exporttoxl, width=15)
        self.btn_export['state'] = 'disable'

        self.mp_check = tk.StringVar()
        self.dwg_check = tk.IntVar()
        self.chk_mp = ttk.Checkbutton(self, text="MP", variable=self.mp_check, onvalue="MP", offvalue="")
        self.chk_dwg = ttk.Checkbutton(self, text="Drawing?", variable=self.dwg_check, onvalue=1, offvalue=0)

        # create table for preview of the BOM list
        # comment out the database source whether production code / test code
        create_db()
        self.columns = ('Category', 'Stock Code', 'Drawing Number', 'Rev', 'Description',
                        'Qty', 'Material', 'Thickness', 'Color')
        self.table_bom = ttk.Treeview(self, columns=self.columns, height=37, show='headings')
        self.table_bom.column('Category', width=100)
        self.table_bom.column('Stock Code', width=75)
        self.table_bom.column('Drawing Number', width=75)
        self.table_bom.column('Rev', width=30, anchor="center")
        self.table_bom.column('Description', width=400)
        self.table_bom.column('Qty', width=50, anchor="center")
        self.table_bom.column('Material', width=50, anchor="center")
        self.table_bom.column('Thickness', width=50, anchor="center")
        self.table_bom.column('Color', width=50, anchor="center")
        self.table_bom.heading('Category', text='Category')
        self.table_bom.heading('Stock Code', text='Stock Code')
        self.table_bom.heading('Drawing Number', text='Drawing No.')
        self.table_bom.heading('Rev', text='Rev')
        self.table_bom.heading('Description', text='Description')
        self.table_bom.heading('Qty', text='Qty')
        self.table_bom.heading('Material', text='Material')
        self.table_bom.heading('Thickness', text='Thk')
        self.table_bom.heading('Color', text='Color')

        self.scrollbar = ttk.Scrollbar(self, orient=tk.VERTICAL, command=self.table_bom.yview)
        self.table_bom.configure(yscrollcommand=self.scrollbar.set)

        # UI for window
        self.up_img = ImageTk.PhotoImage(Image.open('up.png'))
        self.down_img = ImageTk.PhotoImage(Image.open('down.png'))
        self.btn_up = tk.Button(self, image=self.up_img, command=self.move_up,
                                borderwidth=0, repeatdelay=10,
                                width=30, height=70)
        self.btn_down = tk.Button(self, image=self.down_img, command=self.move_down,
                                  borderwidth=0, repeatdelay=10,
                                  width=30, height=70)
        self.menubar = Menu(self)
        self.config(menu=self.menubar)
        self.file_menu = Menu(self.menubar, tearoff=False)
        self.file_menu.add_command(label='Load File', command=self.loadfile)
        self.file_menu.add_command(label='Save File', command=self.savefile)
        self.file_menu.add_separator()
        self.file_menu.add_command(label='Exit', command=self.exit)
        self.rev_menu = Menu(self.menubar, tearoff=False)
        self.rev_menu.add_command(label='Load Revision File', command=self.load_revision)
        self.menubar.add_cascade(label='Menu', menu=self.file_menu, underline=1)
        self.menubar.add_cascade(label='Revision File', menu=self.rev_menu, underline=1)
        self.btn_exit = ttk.Button(self, text="Exit", command=self.exit, width=15)
        self.bind('<Escape>', lambda e: self.exit())
        self.bind('<Return>', lambda  e: self.inserttable())

        # save file and export file variables
        self.todaydate = date.today()
        self.filedate = self.todaydate.strftime("%Y%m%d")

        # widgets placements
        self.lbl_category.grid(row=0, column=0, sticky='w', padx=5)
        self.entry_category.grid(row=0, column=1, columnspan=2, sticky='w', padx=5)
        self.lbl_stockcode.grid(row=1, column=0, sticky='w', padx=5)
        self.entry_stockcode.grid(row=1, column=1, sticky='w', padx=5)
        self.lbl_quantity.grid(row=2, column=0, sticky='w', padx=5)
        self.entry_quantity.grid(row=2, column=1, columnspan=2, sticky='w', padx=5)
        self.chk_mp.grid(row=0, column=3, sticky='w', padx=5)
        self.chk_dwg.grid(row=1, column=3, sticky='w', padx=5)
        self.lbl_message.grid(row=6, column=1, columnspan=2, sticky='w')
        self.lbl_count.grid(row=4, column=1, sticky='e')
        self.lbl_countnumber.grid(row=4, column=2)
        self.lbl_itemqty.grid(row=5, column=1, sticky='e')
        self.lbl_itemqtynumber.grid(row=5, column=2)

        self.btn_table_entry.grid(row=0, column=4, columnspan=2)
        self.btn_remove.grid(row=1, column=4, columnspan=2)
        self.btn_refresh.grid(row=2, column=4, columnspan=2)
        self.btn_exit.grid(row=3, column=4, columnspan=2)
        self.btn_export.grid(row=4, column=4, columnspan=2)

        self.table_bom.grid(row=10, rowspan=40, column=0, columnspan=5, padx=5, pady=5)
        self.scrollbar.grid(row=10, rowspan=40, column=5, sticky='ns')
        self.btn_up.grid(row=10, column=6)
        self.btn_down.grid(row=11, column=6)

    def vlookup(self):
        stockcode = self.entry_stockcode.get().strip()
        list_of_items = query_stockcode(stockcode)
        return list_of_items

    def inserttable(self):
        try:
            if self.entry_stockcode.get() == "":
                self.showmsg.set("No stock code to find")
                self.lbl_message.after(3000, lambda: self.showmsg.set(""))
                return 1
            tablecontent = self.vlookup()[0]
            category = self.entry_category.get().strip()
            quantity = self.entry_quantity.get().strip()

            if tablecontent[4] == "YES":
                self.showmsg.set("Item is obsoleted")
            elif category == "" or quantity == "" or quantity == "0":
                self.showmsg.set("Fill in BOTH category & quantity.")
            elif quantity.isdigit():
                stockcode = tablecontent[1]
                if stockcode not in self.stockcodelist:
                    self.stockcodelist.append(stockcode)

                need_mp = self.mp_check.get()
                need_dwg = self.dwg_check.get()

                if need_dwg == 1:
                    drawingno = need_mp + stockcode.replace('/', '-')
                    self.dwgnumlist.append(drawingno)
                else:
                    drawingno = ""
                rev = ""
                description = tablecontent[2]
                material = self.show_empty_if_none(tablecontent[8])
                thickness = self.show_empty_if_none(tablecontent[9])
                color = self.show_empty_if_none(tablecontent[10])

                # count = int(self.count.get()) + 1
                count = len(self.stockcodelist)
                number = int(self.number.get()) + int(quantity)
                self.count.set(str(count))
                self.number.set(str(number))

                if count == 1:
                    self.table_bom.insert('', tk.END, values=(category, stockcode, drawingno,
                                                              rev, description, quantity,
                                                              material, thickness, color))
                    self.select_latest(0)
                else:
                    for item in self.table_bom.selection():
                        cur_sel = self.table_bom.index(item) + 1
                        self.table_bom.insert('', cur_sel, values=(category, stockcode, drawingno,
                                                                   rev, description, quantity,
                                                                   material, thickness, color))
                        self.select_latest(cur_sel)

                for x in self.table_bom.selection():
                    sel = self.table_bom.index(x)
                    position = (((count - sel) / 2) / count) + 1
                    if sel < 37:
                        self.table_bom.yview_moveto(0)
                    else:
                        self.table_bom.yview_moveto(position)

                self.entry_quantity.delete(0, tk.END)
                self.showmsg.set("Inserting into below table...")
                self.lbl_message.after(3000, lambda: self.showmsg.set(""))
                self.btn_export['state'] = 'disable'
            else:
                self.showmsg.set("Please input quantity as numbers.")
                self.lbl_message.after(3000, lambda: self.showmsg.set(""))
        except IndexError:
            self.showmsg.set("The stockcode is not found in Master Part List")
            self.btn_refresh['state'] = 'active'
            self.lbl_message.after(3000, lambda: self.showmsg.set(""))

    def refresh_database(self):
        self.refresh_status.set("Refreshing...")
        create_db()
        self.btn_refresh.after(3000, lambda: self.refresh_status.set("Refresh DB"))
        self.btn_refresh['state'] = 'disable'

    @staticmethod
    def show_empty_if_none(value):
        return '' if value is None or value == 'None' else value

    def load_revision(self):
        load_rev = askopenfile(parent=self, title="Drawing Revision File",
                               mode='rb', filetypes=[("Excel file", "*.xlsx")])
        if load_rev:
            revwb = openpyxl.load_workbook(load_rev, data_only=True)
            sheet = revwb.active
            maxrow = sheet.max_row
            for i in range(1, maxrow + 1):
                dwgno = sheet.cell(row=i + 1, column=1).value
                rev = sheet.cell(row=i + 1, column=2).value
                self.revdict[dwgno] = rev
        for item in self.table_bom.get_children():
            rowlist = self.table_bom.item(item)["values"]
            dwg_found = rowlist[2]
            if dwg_found != "":
                try:
                    dict_rev = self.revdict[dwg_found]
                    if rowlist[3] == "":
                        table_rev = dict_rev
                        index = self.table_bom.index(item)
                        self.table_bom.delete(item)
                        self.table_bom.insert('', index, values=(rowlist[0], rowlist[1], rowlist[2],
                                                                 table_rev, rowlist[4], rowlist[5],
                                                                 rowlist[6], rowlist[7], rowlist[8]))
                    elif rowlist[3].isdigit():
                        if int(rowlist[3]) < int(dict_rev):
                            table_rev = dict_rev
                            index = self.table_bom.index(item)
                            self.table_bom.delete(item)
                            self.table_bom.insert('', index, values=(rowlist[0], rowlist[1], rowlist[2],
                                                                     table_rev, rowlist[4], rowlist[5],
                                                                     rowlist[6], rowlist[7], rowlist[8]))
                    else:
                        table_rev = rowlist[3]
                        index = self.table_bom.index(item)
                        self.table_bom.delete(item)
                        self.table_bom.insert('', index, values=(rowlist[0], rowlist[1], rowlist[2],
                                                                 table_rev, rowlist[4], rowlist[5],
                                                                 rowlist[6], rowlist[7], rowlist[8]))
                except KeyError:
                    # msgbox.showerror("Revision Not Found", f"{dwg_found} is not found in revision file.")
                    print(f"{dwg_found} is not found in revision file.")

    def select_latest(self, row):
        child = self.table_bom.get_children()
        self.table_bom.focus(child[row])
        self.table_bom.selection_set(child[row])

    def move_up(self):
        for item in self.table_bom.selection():
            self.table_bom.move(item, self.table_bom.parent(item), self.table_bom.index(item) - 1)

    def move_down(self):
        for item in reversed(self.table_bom.selection()):
            self.table_bom.move(item, self.table_bom.parent(item), self.table_bom.index(item) + 1)

    def exporttoxl(self):
        if self.status == 1:
            inputname = self.inputname
            filename = inputname + self.filedate + '.xlsx'
        else:
            filepath = self.inputname
            maxchar = 0
            for i, char in enumerate(filepath, start=1):
                if char == "/":
                    if i > maxchar:
                        maxchar = i
            filename = filepath[maxchar:]
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb["Sheet1"]
        ws.insert_rows(1)
        for i, item in enumerate(self.columns, start=1):
            ws.cell(row=1, column=i).value = item
        wb.save(filename)
        excelpath = Path(filename).resolve()
        os.system(f'start excel.exe "{excelpath}')

    def removeitem(self):
        for item in self.table_bom.selection():
            cur_sel = self.table_bom.index(item)
            qty = int(self.table_bom.item(item)["values"][5])
            self.table_bom.delete(item)
            if len(self.table_bom.get_children()) > 0:
                self.select_latest(cur_sel - 1)
            count = int(self.count.get()) - 1
            number = int(self.number.get())
            self.count.set(str(count))
            self.number.set(str(number - qty))

    def loadfile(self):
        load_excel = askopenfile(parent=self, title="Choose your project to continue...",
                                 mode='rb', filetypes=[("Excel file", "*.xlsx")])

        if load_excel:
            workbook = openpyxl.load_workbook(load_excel, data_only=True)
            sheet = workbook.active
            maxrow = sheet.max_row
            self.inputname = load_excel.name
            qty = 0
            for i in range(1, maxrow + 1):
                listrow = []
                for item in sheet[i]:
                    checked_item = self.show_empty_if_none(item.value)
                    listrow.append(checked_item)
                self.table_bom.insert('', tk.END, values=listrow)

            for cell in sheet['F']:
                qty = int(cell.value) + qty
            self.number.set(qty)
            self.select_latest(-1)

            for cell in sheet['B']:
                if cell.value not in self.stockcodelist:
                    self.stockcodelist.append(cell.value)
            self.count.set(str(len(self.stockcodelist)))

            for cell in sheet['C']:
                if cell.value is not None:
                    self.dwgnumlist.append(cell.value)

    def savefile(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
        if self.status == 1:
            inputname = self.inputname
            filename = self.filedate + '_' + inputname + '.xlsx'
            workbook.save(filename)
        else:
            filepath = self.inputname
            maxchar = 0
            for i, char in enumerate(filepath, start=1):
                if char == "/":
                    if i > maxchar:
                        maxchar = i
            filename = filepath[maxchar:]
        i = 1
        for item in self.table_bom.get_children():
            j = 1
            rowlist = self.table_bom.item(item)["values"]
            for r in rowlist:
                sheet.cell(row=i, column=j).value = r
                j += 1
            i += 1
        workbook.save(filename)
        self.btn_export['state'] = 'active'

    def exit(self):
        if msgbox.askyesno("Exit", "Are you sure to exit?"):
            self.destroy()


class App(tk.Tk):
    inputname = ""

    def __init__(self):
        super().__init__()

        self.title("Welcome")
        self.btn_new = ttk.Button(self, text="New", command=self.new_entry, width=100)
        self.btn_load = ttk.Button(self, text="Load", command=self.load_entry, width=100)

        self.btn_new.grid(row=0, column=0, padx=25, pady=20)
        self.btn_load.grid(row=1, column=0, padx=25, pady=20)

    def new_entry(self):
        self.inputname = simpledialog.askstring(title="Save As", prompt="File Name: ")
        work_window = MainWindow(self)
        work_window.inputname = self.inputname
        if work_window.inputname.strip() != "":
            work_window.status = 1
            work_window.grab_set()
        else:
            msgbox.showwarning("Warning!", "No title given!")
            sys.exit()

    def load_entry(self):
        work_window = MainWindow(self)
        work_window.grab_set()
        work_window.loadfile()


if __name__ == "__main__":
    main = App()
    main.mainloop()
